import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from html import escape

def excel_to_html(file_path):
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active
    html_table = "<table border='1' style='border-collapse: collapse;'>"
    merged_ranges = sheet.merged_cells.ranges

    color_mapping = {
        'N1-PROD': 'background-color: #ADD8E6',
        'KPI\'S': 'background-color: #F0F0F0',
        'ETCD INFO': 'background-color: #D3D3D3',
        'SERVICES': 'background-color: #D3D3D3',
        'PODS': 'background-color: #D3D3D3',
        'NODES': 'background-color: #D3D3D3',
        'DEPLOYMENT': 'background-color: #D3D3D3',
        'ALERTS': 'background-color: #D3D3D3',
        'BACKUP': 'background-color: #D3D3D3',
    }

    green_values = ['UP', 'Yes', '100%']
    
    for row in sheet.iter_rows():
        html_table += "<tr>"
        for cell in row:
            rowspan, colspan = 1, 1
            is_merged_start = False
            for merged_range in merged_ranges:
                if cell.coordinate in merged_range:
                    if cell.coordinate == merged_range.start_cell.coordinate:
                        is_merged_start = True
                        rowspan = merged_range.size['rows']
                        colspan = merged_range.size['columns']
                    else:
                        is_merged_start = False
                        break

            if not is_merged_start and any(cell.coordinate in merged_range for merged_range in merged_ranges):
                continue

            # Handle percentage values
            if cell.number_format and '%' in cell.number_format:
                value = f"{cell.value * 100}%" if cell.value is not None else ""
            else:
                value = str(cell.value) if cell.value is not None else ""
            
            escaped_value = escape(value)

            styles = ['padding: 5px']

            if value in color_mapping:
                styles.append(color_mapping[value])
            elif value in green_values:
                styles.append('background-color: #90EE90')
                
            # elif isinstance(cell.value, (int,float)):
            #     try:
            #         numeric_value = float (cell.value)
            #         if numeric_value < 5:
            #             styles.append('background-color: #90EE90')  # Green
            #         else:
            #             styles.append('background-color: #FFB6B6')  # Red
            #     except ValueError:
            #         pass
                
            elif value == '4.9':
                try:
                    numeric_value = float(value)
                    if numeric_value < 5:
                        styles.append('background-color: #90EE90')
                    else:
                        styles.append('background-color: #FFB6B6')
                except ValueError:
                    pass
            
            else:
                if cell.fill and isinstance(cell.fill, PatternFill) and cell.fill.start_color:
                    bg_color = cell.fill.start_color.rgb
                    if bg_color and bg_color != '00000000':
                        if isinstance(bg_color, str):
                            if bg_color.startswith('FF'):
                                bg_color = bg_color[2:]
                            styles.append(f"background-color: #{bg_color}")

            if cell.font and cell.font.bold:
                styles.append("font-weight: bold")

            if cell.alignment:
                if cell.alignment.horizontal:
                    styles.append(f"text-align: {cell.alignment.horizontal}")
                if cell.alignment.vertical:
                    styles.append(f"vertical-align: {cell.alignment.vertical}")

            style_string = f"style='{'; '.join(styles)}'" if styles else ""
            
            if rowspan > 1 or colspan > 1:
                html_table += f"<td {style_string} rowspan='{rowspan}' colspan='{colspan}'>{escaped_value}</td>"
            else:
                html_table += f"<td {style_string}>{escaped_value}</td>"

        html_table += "</tr>"

    html_table += "</table>"
    return html_table

# <p>Hello, this is test email of daily SLA report</p>

# Email configuration
sender_email = "amitesh.singh@coredge.io"
receiver_emails = ["amitesh.singh@coredge.io"]
                #    "gyan.bhatnagar@coredge.io",
                #    "rajat@coredge.io", 
                #    "raunak@coredge.io"
# cc_emails = ["cellstack-support@coredge.io"]
cc_emails = ["amitesh.singh@coredge.io"]
email_password = "xyz"
subject = "Daily Monitoring Report"

input_excel_file = "Daily_monitoring_report.xlsx"

html_table = excel_to_html(input_excel_file)

message = MIMEMultipart()
message["From"] = sender_email
message["To"] = ", ".join(receiver_emails)
message["Cc"] = ", ".join(cc_emails)
message["Subject"] = subject

body = f"""
<p>Hello all,</p>
<p>This is today's SLA report data.</p>
{html_table}
"""
message.attach(MIMEText(body, "html"))

with open(input_excel_file, "rb") as attachment:
    part = MIMEBase("application", "octet-stream")
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header(
        "Content-Disposition",
        f"attachment; filename={input_excel_file.split('/')[-1]}",
    )
    message.attach(part)

try:
    all_recipients = receiver_emails + cc_emails
    with smtplib.SMTP("smtp.office365.com", 587) as server:
        server.starttls()
        server.login(sender_email, email_password)
        server.sendmail(sender_email, all_recipients, message.as_string())
    print("Email sent successfully to all recipients!")
except smtplib.SMTPAuthenticationError as e:
    print("Authentication Error: Check your username and password or App Password settings.")
    print(e)
except Exception as e:
    print("Error sending email:", e)
